# -*- coding:utf-8 -*-
# --------------------------------------------------------
# Copyright (C), 2016-2021, lizhe, All rights reserved
# --------------------------------------------------------
# @Name:        excel_utils.py
# @Author:      lizhe
# @Created:     2021/12/25 - 22:18
# --------------------------------------------------------
from typing import Sequence, Optional, Union, Dict
from automotive.utils.common.interfaces import BaseExcelUtils, wb, sht, cell_range
import xlwings as xw
import openpyxl
from .common.enums import ExcelEnum
from ..logger.logger import logger
from xlwings import Sheet, Book, Range
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment


class OpenpyxlExcelUtils(BaseExcelUtils):

    def create_workbook(self) -> Workbook:
        return Workbook()

    def create_sheet(self, workbook: Workbook, sheet_name: str) -> Worksheet:
        return workbook.create_sheet(sheet_name)

    def open_workbook(self, file: str) -> Workbook:
        return openpyxl.load_workbook(file)

    def close_workbook(self, workbook: Workbook):
        workbook.close()

    def get_sheets(self, workbook: Workbook) -> Sequence:
        return workbook.worksheets

    def get_sheet_dict(self, workbook: wb) -> Dict[str, sht]:
        sheet_dict = dict()
        sheets = self.get_sheets(workbook)
        for sheet in sheets:
            sheet_dict[sheet.title] = sheet
        return sheet_dict

    def get_sheet(self, workbook: Workbook, sheet_name: str) -> Worksheet:
        return workbook.get_sheet_by_name(sheet_name)

    def get_sheet_name(self, sheet: sht) -> str:
        return sheet.title

    def copy_sheet(self, workbook: Workbook, origin_sheet_name: str, target_sheet_name: str) -> Worksheet:
        origin_sheet = self.get_sheet(workbook, origin_sheet_name)
        target_sheet = workbook.copy_worksheet(origin_sheet)
        target_sheet.title = target_sheet_name
        return target_sheet

    def delete_sheet(self, workbook: Workbook, sheet_name: str):
        sheet = self.get_sheet(workbook, sheet_name)
        workbook.remove(sheet)

    def get_max_rows(self, sheet: sht) -> int:
        return sheet.max_row

    def get_max_columns(self, sheet: sht) -> int:
        return sheet.max_column

    def get_sheet_contents(self, sheet: Worksheet, start_row: int = 1) -> Sequence:
        contents = []
        max_row = self.get_max_rows(sheet) + 1
        max_column = self.get_max_columns(sheet) + 1
        logger.debug(f"max_row = {max_row}")
        logger.debug(f"max_column = {max_column}")
        for row in range(start_row, max_row):
            lines = []
            for column in range(1, max_column):
                logger.debug(f"get cell[{row},{column}]")
                line = self.get_cell_value(sheet, row, column)
                # 读不到内容则为空字符串
                line = line if line else ""
                lines.append(line)
            logger.debug(f"lines = {lines}")
            contents.append(lines)
        return contents

    def set_sheet_contents(self, sheet: Worksheet, contents: Sequence, start_row: int = 1, border: bool = False):
        for row, lines in enumerate(contents):
            logger.info(f"write row {row + start_row}.....")
            for column, value in enumerate(lines):
                row_index = start_row + row
                column_index = column + 1
                logger.debug(f"write cell[{row_index}, {column_index}] = {value}")
                self.set_cell_value(sheet, row_index, column_index, value, border)

    def get_cell_value(self, sheet: Worksheet, row_index: int, column_index: Union[str, int]) -> str:
        # 根据列名获取序号
        if isinstance(column_index, str):
            column_index = self._get_column_index(column_index)
        value = sheet.cell(row_index, column_index).value
        logger.trace(f"cell[{row_index}, {column_index}] = {value}")
        return value

    def set_cell_value(self, sheet: Worksheet, row_index: int, column_index: Union[str, int], value: str,
                       border: bool = False):
        # 根据列名获取序号
        if isinstance(column_index, str):
            column_index = self._get_column_index(column_index)
        cell = sheet.cell(row_index, column_index)
        if border:
            self.set_border(cell)
        cell.value = value

    def save_workbook(self, file: str, workbook: Workbook):
        workbook.save(filename=file)

    def set_border(self, cell: cell_range):
        cell.border = Border(left=Side(border_style="thin", color="000000"),
                             right=Side(border_style="thin", color="000000"),
                             top=Side(border_style="thin", color="000000"),
                             bottom=Side(border_style="thin", color="000000"))
        # 居中显示 + 自动换行
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    @staticmethod
    def split_merge_cell(workbook: Workbook, worksheet: Worksheet, xlsx_path: str, save_path: str):
        """
        拆分合并了的单元格，存储到新的xlsx文件中.仅支持一列的合并单元格，一行单元格目前不支持
        :param worksheet:
        :param workbook:
        :param xlsx_path: 合并单元格的excel
        :param save_path: 拆分了单元格的excel
        :return:
        """
        if save_path == xlsx_path:
            raise RuntimeWarning("拆分路径不可以跟xlsx文件路径一样\n")

        # m_list合并单元格的位置信息，可迭代对象（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是excel坐标信息
        m_list = worksheet.merged_cells
        if m_list:
            cr = []
            for m_area in m_list:
                # 合并单元格的起始行坐标、终止行坐标。。。。，
                r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
                # 纵向合并单元格的位置信息提取出
                if r2 - r1 > 0:
                    cr.append((r1, r2, c1, c2))
            # 这里注意需要把合并单元格的信息提取出再拆分
            for r in cr:
                worksheet.unmerge_cells(start_row=r[0], end_row=r[1],
                                        start_column=r[2], end_column=r[3])
                for i in range(r[1] - r[0] + 1):
                    for j in range(r[3] - r[2] + 1):
                        worksheet.cell(row=r[0] + i, column=r[2] + j, value=worksheet.cell(r[0], r[2]).value)
            workbook.save(save_path)


class XlwingsExcelUtils(BaseExcelUtils):

    def __init__(self):
        self.__app = xw.App(visible=False, add_book=False)
        self.__app.display_alerts = False
        self.__app.screen_updating = False

    @staticmethod
    def __compare_sheets(before_sheets: Sequence[str], after_sheets: Sequence[str]):
        for sheet in after_sheets:
            if sheet not in before_sheets:
                return sheet
        raise RuntimeError("copy failed")

    def __get_sheet_names(self, workbook: Book):
        worksheets = self.get_sheets(workbook)
        names = []
        for worksheet in worksheets:
            names.append(worksheet.name)
        return names

    def create_workbook(self) -> Book:
        return self.__app.books.add()

    def create_sheet(self, workbook: Book, sheet_name: str) -> Sheet:
        return workbook.sheets.add(sheet_name)

    def open_workbook(self, file: str) -> Book:
        return self.__app.books.open(file)
        # return xw.Book(file)

    def close_workbook(self, workbook: Book):
        self.__app.kill()

    def get_sheet_dict(self, workbook: wb) -> Dict[str, sht]:
        sheet_dict = dict()
        worksheets = self.get_sheets(workbook)
        for worksheet in worksheets:
            sheet_dict[worksheet.name] = worksheet
        return sheet_dict

    def get_sheets(self, workbook: Book) -> Sequence[Sheet]:
        return workbook.sheets

    def get_sheet(self, workbook: Book, sheet: str) -> Sheet:
        return workbook.sheets[sheet]

    def get_sheet_name(self, sheet: sht) -> str:
        return sheet.name

    def copy_sheet(self, workbook: Book, origin_sheet: str, target_sheet: str) -> Sheet:
        # 要被copy的sheet
        sheet = workbook.sheets[origin_sheet]
        # 获取当前所有sheet的名字
        before = self.__get_sheet_names(workbook)
        # 复制sheet
        sheet.api.Copy(Before=sheet.api)
        after = self.__get_sheet_names(workbook)
        sheet_name = self.__compare_sheets(before, after)
        new_sheet = workbook.sheets[sheet_name]
        new_sheet.name = target_sheet
        return new_sheet

    def delete_sheet(self, workbook: Book, sheet: Optional[str]):
        self.get_sheet(workbook, sheet).delete()

    def get_max_rows(self, sheet: sht) -> int:
        return sheet.used_range.last_cell.row

    def get_max_columns(self, sheet: sht) -> int:
        return sheet.used_range.last_cell.column

    def get_sheet_contents(self, sheet: Sheet, start_row: int = 1) -> Sequence:
        contents = []
        max_row = self.get_max_rows(sheet) + 1
        max_column = self.get_max_columns(sheet) + 1
        for row in range(start_row, max_row + 1):
            lines = []
            for column in range(max_column):
                line = self.get_cell_value(sheet, row, column + 1)
                # 读不到内容则为空字符串
                line = line if line else ""
                lines.append(line)
            logger.debug(f"lines = {lines}")
            contents.append(lines)
        return contents

    def set_sheet_contents(self, sheet: Sheet, contents: Sequence, start_row: int = 1, border: bool = False):
        for row, lines in enumerate(contents):
            logger.info(f"write row {row + start_row}.....")
            for column, value in enumerate(lines):
                row_index = start_row + row
                column_index = column + 1
                logger.debug(f"write cell[{row_index}, {column_index}] = {value}")
                self.set_cell_value(sheet, row_index, column_index, value, border)

    def get_cell_value(self, sheet: Sheet, row_index: int, column_index: Union[str, int]) -> str:
        # 传递的是序号则获取列名
        if isinstance(column_index, int):
            column_index = self._get_column_name(column_index)
        logger.debug(f"column_name = {column_index}, row_index = {row_index}")
        value = sheet.range(f"{column_index}{row_index}").value
        logger.debug(f"cell[{row_index}, {column_index}] = {value}")
        return value

    def set_cell_value(self, sheet: Sheet, row_index: int, column_index: Union[str, int], value: str,
                       border: bool = False):
        # 传递的是序号则获取列名
        if isinstance(row_index, int):
            column_index = self._get_column_name(column_index)
        cell = sheet.range(f"{column_index}{row_index}")
        if border:
            self.set_border(cell)
        cell.value = value

    def save_workbook(self, file: str, workbook: Book):
        workbook.save(file)

    def set_border(self, cell: Range):
        borders = 7, 8, 9, 10
        # 底部边框 9 左边框 7 顶部框 8 右边框 10
        for border in borders:
            cell.api.Borders(border).LineStyle = 1
        # 居中显示
        cell.api.HorizontalAlignment = -4108
        # 自动换行
        cell.api.WrapText = True

    @staticmethod
    def del_row(sheet: Sheet, row_index: int):
        sheet.api.Rows(row_index).Delete()


class ExcelUtils(BaseExcelUtils):

    def __init__(self, type_: Union[ExcelEnum, str] = ExcelEnum.OPENPYXL):
        if isinstance(type_, str):
            type_ = ExcelEnum.from_name(type_)
        if type_ == ExcelEnum.XLWINGS:
            self.__utils = XlwingsExcelUtils()
        elif type_ == ExcelEnum.OPENPYXL:
            self.__utils = OpenpyxlExcelUtils()
        else:
            raise RuntimeError(f"{type_.value} not support")

    def create_workbook(self) -> wb:
        return self.__utils.create_workbook()

    def create_sheet(self, workbook: wb, sheet_name: str) -> sht:
        return self.__utils.create_sheet(workbook, sheet_name)

    def open_workbook(self, file: str) -> wb:
        return self.__utils.open_workbook(file)

    def get_sheet_dict(self, workbook: wb) -> Dict[str, sht]:
        return self.__utils.get_sheet_dict(workbook)

    def get_sheets(self, workbook: wb) -> Sequence[sht]:
        return self.__utils.get_sheets(workbook)

    def get_sheet(self, workbook: wb, sheet_name: str) -> sht:
        return self.__utils.get_sheet(workbook, sheet_name)

    def get_sheet_name(self, sheet: sht) -> str:
        return self.__utils.get_sheet_name(sheet)

    def copy_sheet(self, workbook: wb, origin_sheet: str, target_sheet: str) -> sht:
        return self.__utils.copy_sheet(workbook, origin_sheet, target_sheet)

    def delete_sheet(self, workbook: wb, sheet: str):
        self.__utils.delete_sheet(workbook, sheet)

    def get_sheet_contents(self, sheet: sht, start_row: int = 1) -> Sequence:
        return self.__utils.get_sheet_contents(sheet, start_row)

    def set_sheet_contents(self, sheet: sht, contents: Sequence, start_row: int = 1, border: bool = False):
        self.__utils.set_sheet_contents(sheet, contents, start_row, border)

    def get_max_rows(self, sheet: sht) -> int:
        return self.__utils.get_max_rows(sheet)

    def get_max_columns(self, sheet: sht) -> int:
        return self.__utils.get_max_columns(sheet)

    def get_cell_value(self, sheet: sht, row_index: int, column_index: Union[str, int]) -> str:
        return self.__utils.get_cell_value(sheet, row_index, column_index)

    def set_cell_value(self, sheet: sht, row_index: int, column_index: Union[str, int], value: str,
                       border: bool = False):
        self.__utils.set_cell_value(sheet, row_index, column_index, value, border)

    def save_workbook(self, file: str, workbook: wb):
        self.__utils.save_workbook(file, workbook)

    def close_workbook(self, workbook: wb):
        self.__utils.close_workbook(workbook)

    def set_border(self, cell: cell_range):
        self.__utils.set_border(cell)
