# -*- coding:utf-8 -*-
# --------------------------------------------------------
# Copyright (C), 2016-2021, lizhe, All rights reserved
# --------------------------------------------------------
# @Name:        reader.py
# @Author:      lizhe
# @Created:     2021/12/13 - 22:25
# --------------------------------------------------------
import os
from typing import List, Dict, Any

from automotive.application.common.constants import GuiConfig
from automotive.application.common.interfaces import BaseConfigReader
from automotive.logger.logger import logger
from automotive.application.common.enums import GuiButtonTypeEnum, ExcelReadEnum

try:
    import xlwings as xw
except ModuleNotFoundError:
    os.system("pip install xlwings")
finally:
    import xlwings as xw
    from xlwings import Sheet as xw_Sheet
    from xlwings import Book

try:
    import xlrd
except ModuleNotFoundError:
    os.system("pip install xlrd")
finally:
    import xlrd
    from xlrd.sheet import Sheet as xlrd_Sheet

try:
    import openpyxl
except ModuleNotFoundError:
    os.system("pip install openpyxl")
finally:
    import openpyxl
    from openpyxl import worksheet

check_buttons = GuiButtonTypeEnum.CHECK_BUTTON.value[1]
thread_buttons = GuiButtonTypeEnum.EVENT_CHECK_BUTTON.value[1]
comboxs = GuiButtonTypeEnum.COMBOX_BUTTON.value[1]
entries = GuiButtonTypeEnum.INPUT_BUTTON.value[1]
buttons = GuiButtonTypeEnum.EVENT_BUTTON.value[1]


class XlwingsConfigReader(BaseConfigReader):

    def read_from_file(self, file: str) -> Dict[str, Dict[str, Any]]:
        result = dict()
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        wb = app.books.open(file)
        sheet = wb.sheets["sheet"]
        configs = self.__parse(sheet)
        # 先区分tab
        tabs = self._split_tabs(configs)
        for tab, values in tabs.items():
            tab_config = dict()
            # 按照按钮类型来分割
            typed_configs = self._split_type(values)
            # 处理 buttons
            tab_config[check_buttons] = self._handle_check_buttons(typed_configs[GuiButtonTypeEnum.CHECK_BUTTON])
            # 处理 flash_buttons
            tab_config[thread_buttons] = self._handle_event_buttons(
                typed_configs[GuiButtonTypeEnum.EVENT_CHECK_BUTTON])
            # 处理 comboxs
            tab_config[comboxs] = self._handle_combox(typed_configs[GuiButtonTypeEnum.COMBOX_BUTTON])
            # 处理 entries
            tab_config[entries] = self._handle_entries(typed_configs[GuiButtonTypeEnum.INPUT_BUTTON])
            # 处理 buttons
            tab_config[buttons] = self._handle_buttons(typed_configs[GuiButtonTypeEnum.EVENT_BUTTON])
            result[tab] = tab_config
        wb.close()
        app.quit()
        try:
            app.kill()
        except AttributeError:
            logger.debug("app kill fail")
        return result

    def __parse(self, sheet: xw_Sheet) -> List[GuiConfig]:
        configs = []
        max_row = sheet.used_range.last_cell.row
        for i in range(2, max_row + 1):
            config = GuiConfig()
            config.name = sheet.range(f"A{i}").value
            config.text_name = sheet.range(f"B{i}").value
            config.button_type = GuiButtonTypeEnum.from_name(sheet.range(f"C{i}").value)
            column_d = sheet.range(f"D{i}").value
            config.selected = self._parse_actions(column_d) if column_d else None
            column_e = sheet.range(f"E{i}").value
            config.unselected = self._parse_actions(column_e) if column_e else None
            config.items = sheet.range(f"F{i}").value
            column_g = sheet.range(f"G{i}").value
            config.actions = self._parse_actions(column_g) if column_g else None
            config.tab_name = sheet.range(f"H{i}").value
            configs.append(config)
            logger.debug(f"config = {config}")
        return configs


class OpenpyxlConfigReader(BaseConfigReader):

    def read_from_file(self, file: str) -> Dict[str, Dict[str, Any]]:
        result = dict()
        wb = openpyxl.load_workbook(file)
        sheet = wb["sheet"]
        configs = self.__parse(sheet)
        # 先区分tab
        tabs = self._split_tabs(configs)
        for tab, values in tabs.items():
            logger.debug(f"tab name is {tab}")
            tab_config = dict()
            # 按照按钮类型来分割
            typed_configs = self._split_type(values)
            # 处理 buttons
            tab_config[check_buttons] = self._handle_check_buttons(typed_configs[GuiButtonTypeEnum.CHECK_BUTTON])
            # 处理 flash_buttons
            tab_config[thread_buttons] = self._handle_event_buttons(typed_configs[GuiButtonTypeEnum.EVENT_CHECK_BUTTON])
            # 处理 comboxs
            tab_config[comboxs] = self._handle_combox(typed_configs[GuiButtonTypeEnum.COMBOX_BUTTON])
            # 处理 entries
            tab_config[entries] = self._handle_entries(typed_configs[GuiButtonTypeEnum.INPUT_BUTTON])
            # 处理 buttons
            tab_config[buttons] = self._handle_buttons(typed_configs[GuiButtonTypeEnum.EVENT_BUTTON])
            result[tab] = tab_config
        return result

    def __parse(self, sheet: worksheet):
        configs = []
        max_row = sheet.max_row
        logger.debug(f"max_row = {max_row}")
        for i in range(2, max_row + 1):
            config = GuiConfig()
            config.name = sheet.cell(i, 1).value
            config.text_name = sheet.cell(i, 2).value
            config.button_type = GuiButtonTypeEnum.from_name(sheet.cell(i, 3).value)
            column_d = sheet.cell(i, 4).value
            config.selected = self._parse_actions(column_d) if column_d else None
            column_e = sheet.cell(i, 5).value
            config.unselected = self._parse_actions(column_e) if column_e else None
            config.items = sheet.cell(i, 6).value
            column_g = sheet.cell(i, 7).value
            config.actions = self._parse_actions(column_g) if column_g else None
            config.tab_name = sheet.cell(i, 8).value
            configs.append(config)
            logger.debug(f"config = {config}")
        return configs


class XlrdConfigReader(BaseConfigReader):

    def read_from_file(self, file: str) -> Dict[str, Dict[str, Any]]:
        result = dict()
        data = xlrd.open_workbook(file)
        sheet = data.sheet_by_name("sheet")
        configs = self.__parse(sheet)
        # 先区分tab
        tabs = self._split_tabs(configs)
        for tab, values in tabs.items():
            tab_config = dict()
            # 按照按钮类型来分割
            typed_configs = self._split_type(values)
            # 处理 buttons
            tab_config[check_buttons] = self._handle_check_buttons(typed_configs[GuiButtonTypeEnum.CHECK_BUTTON])
            # 处理 flash_buttons
            tab_config[thread_buttons] = self._handle_event_buttons(typed_configs[GuiButtonTypeEnum.EVENT_CHECK_BUTTON])
            # 处理 comboxs
            tab_config[comboxs] = self._handle_combox(typed_configs[GuiButtonTypeEnum.COMBOX_BUTTON])
            # 处理 entries
            tab_config[entries] = self._handle_entries(typed_configs[GuiButtonTypeEnum.INPUT_BUTTON])
            # 处理 buttons
            tab_config[buttons] = self._handle_buttons(typed_configs[GuiButtonTypeEnum.EVENT_BUTTON])
            result[tab] = tab_config
        return result

    def __parse(self, sheet: xlrd_Sheet):
        configs = []
        max_row = sheet.nrows
        for i in range(2, max_row + 1):
            config = GuiConfig()
            config.name = sheet.cell_value(i, 0)
            config.text_name = sheet.cell_value(i, 1)
            config.button_type = GuiButtonTypeEnum.from_name(sheet.cell_value(i, 2))
            column_d = sheet.cell_value(i, 3)
            config.selected = self._parse_actions(column_d) if column_d else None
            column_e = sheet.cell_value(i, 4)
            config.unselected = self._parse_actions(column_e) if column_e else None
            config.items = sheet.cell_value(i, 5)
            column_g = sheet.cell_value(i, 6).value
            config.actions = self._parse_actions(column_g) if column_g else None
            config.tab_name = sheet.cell_value(i, 7).value
            configs.append(config)
            logger.debug(f"config = {config}")
        return configs


class ConfigService(object):

    def __init__(self, type_: ExcelReadEnum = ExcelReadEnum.OPENPYXL):
        if type_ == ExcelReadEnum.XLRD:
            self.__reader = XlrdConfigReader()
        elif type_ == ExcelReadEnum.XLWINGS:
            self.__reader = XlwingsConfigReader()
        elif type_ == ExcelReadEnum.OPENPYXL:
            self.__reader = OpenpyxlConfigReader()

    def read_from_file(self, file: str) -> Dict[str, Dict[str, Any]]:
        return self.__reader.read_from_file(file)
